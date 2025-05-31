from abc import ABC
from dataclasses import dataclass, field
import enum
import logging
from typing import List, Optional, Tuple, Union
from PIL import Image
import pymupdf

from openpyxl import Workbook
from openpyxl.utils import get_column_letter 
from openpyxl.styles import Alignment, Border, Side 


@dataclass
class BBox:
    x1: int
    y1: int
    x2: int
    y2: int

    @property
    def coords(self) -> Tuple[int, int, int, int]:
        """Вернуть все координаты как кортеж."""
        return self.x1, self.y1, self.x2, self.y2

    @coords.setter
    def coords(self, vals: Tuple[int, int, int, int]):
        """Установить сразу все координаты из кортежа."""
        self.x1, self.y1, self.x2, self.y2 = vals

    @property
    def width(self) -> int:
        """Ширина bbox."""
        return self.x2 - self.x1

    @property
    def height(self) -> int:
        """Высота bbox."""
        return self.y2 - self.y1

    def padding(self, pix: int = 5) -> "BBox":
        return BBox(
            x1=self.x1 - pix,
            y1=self.y1 - pix,
            x2=self.x2 + pix,
            y2=self.y2 + pix,
        )
    
    def contains(self, other: 'BBox') -> bool:
        return (self.x1 <= other.x1 and
                self.y1 <= other.y1 and
                self.x2 >= other.x2 and
                self.y2 >= other.y2)

    @classmethod
    def from_rect(
        cls,
        rect: Union[
            "pymupdf.Rect",      # если у вас есть real Rect
            Tuple[float, float, float, float]  # или кортеж (x0,y0,x1,y1)
        ],
        sx: float = 1.0,
        sy: float = 1.0
    ) -> "BBox":
        """Создать BBox из pymupdf.Rect или из кортежа (x0, y0, x1, y1) с учётом масштабов."""
        # разбираем вход
        if hasattr(rect, "x0"):
            x0, y0, x1, y1 = rect.x0, rect.y0, rect.x1, rect.y1
        else:
            x0, y0, x1, y1 = rect  # ожидаем кортеж из 4-х чисел

        return cls(
            x1=int(x0 * sx),
            y1=int(y0 * sy),
            x2=int(x1 * sx),
            y2=int(y1 * sy),
        )

class InsertionPosition(enum.Enum):
    TOP = "top"  # Сверху
    BOTTOM = "bottom" # Снизу
    LEFT = "left"
    RIGHT = "right"

@dataclass
class Cell:
    bbox: BBox
    row: int
    col: int
    colspan: int
    rowspan: int
    text: str = None
    blobs: List[BBox] = field(default_factory=list)
    original_page_num: Optional[int] = None

    @property
    def has_text(self) -> bool:
        """Проверяет, содержит ли ячейка текст (на основе атрибута text или blobs)."""
        if self.text and self.text.strip():
            return True
        return len(self.blobs) > 0
    
    @property
    def free_space_ratio(self) -> float:
        """
        Оценивает долю свободного пространства внутри bbox ячейки, не занятого blobs.
        Возвращает 1.0, если bbox ячейки имеет нулевую площадь.
        """
        cell_area = self.bbox.width * self.bbox.height
        if cell_area == 0:
            return 1.0  # Если площадь ячейки 0, считаем ее полностью свободной или полностью занятой в зависимости от контекста. 1.0 означает нет места для нового.
                        # Или можно вернуть 0.0 если blobs тоже нет, что означает "нет контента, нет места"

        occupied_area_by_blobs = 0
        for blob in self.blobs:
            occupied_area_by_blobs += blob.width * blob.height
        
        # Ограничиваем occupied_area, чтобы она не превышала cell_area
        occupied_area_by_blobs = min(occupied_area_by_blobs, float(cell_area))

        free_area = cell_area - occupied_area_by_blobs
        return free_area / cell_area

    def get_insertion_area(
        self,
        position: InsertionPosition,
        min_height: int = 10, # Минимальная высота для области вставки
        padding: int = 2      # Отступ от краев основного bbox ячейки
    ) -> Optional[BBox]:
        """
        Вычисляет прямоугольную область внутри ячейки, подходящую для вставки нового текста.
        Область будет занимать ширину ячейки (за вычетом отступов).

        Args:
            position: Член перечисления InsertionPosition (TOP или BOTTOM).
            min_height: Минимально необходимая высота для области вставки.
            padding: Отступ от краев основного bbox ячейки.

        Returns:
            Объект BBox, представляющий доступную область, или None, если подходящая область не найдена.
        """
        if self.bbox.width <= 2 * padding or self.bbox.height <= 2 * padding:
            return None # Ячейка слишком мала для отступов

        # Эффективные границы ячейки после применения отступов
        eff_cell_x1 = self.bbox.x1 + padding
        eff_cell_y1 = self.bbox.y1 + padding
        eff_cell_x2 = self.bbox.x2 - padding
        eff_cell_y2 = self.bbox.y2 - padding

        if eff_cell_x1 >= eff_cell_x2 or eff_cell_y1 >= eff_cell_y2:
            return None # Нет места после применения отступов

        insert_x1 = eff_cell_x1
        insert_x2 = eff_cell_x2
        insert_y1 = -1
        insert_y2 = -1

        # Фильтруем blobs, которые находятся по горизонтали в пределах эффективной ширины ячейки
        relevant_blobs = [
            b for b in self.blobs if b.x1 < eff_cell_x2 and b.x2 > eff_cell_x1
        ]

        if position == InsertionPosition.TOP:
            insert_y1 = eff_cell_y1 
            limit_y = eff_cell_y2 # По умолчанию, предел - это нижняя граница ячейки с отступом
            if relevant_blobs:
                # Ищем самый верхний край существующих blobs, чтобы вставить текст над ними
                blob_tops = [b.y1 for b in relevant_blobs if b.y1 >= eff_cell_y1] 
                if blob_tops:
                    limit_y = min(min(blob_tops) - 1, eff_cell_y2) # -1 для небольшого зазора
            insert_y2 = limit_y

        elif position == InsertionPosition.BOTTOM:
            insert_y2 = eff_cell_y2
            limit_y = eff_cell_y1 # По умолчанию, предел - это верхняя граница ячейки с отступом
            if relevant_blobs:
                # Ищем самый нижний край существующих blobs, чтобы вставить текст под ними
                blob_bottoms = [b.y2 for b in relevant_blobs if b.y2 <= eff_cell_y2]
                if blob_bottoms:
                    limit_y = max(max(blob_bottoms) + 1, eff_cell_y1) # +1 для небольшого зазора
            insert_y1 = limit_y
        else:
            # Это можно расширить для других позиций, таких как LEFT, RIGHT и т.д.
            raise NotImplementedError(f"Позиция для вставки {position} еще не поддерживается.")

        # Проверяем рассчитанную область
        if insert_y1 != -1 and insert_y2 != -1 and \
        insert_x1 < insert_x2 and insert_y1 < insert_y2 and \
        (insert_y2 - insert_y1) >= min_height:
            return BBox(x1=insert_x1, y1=insert_y1, x2=insert_x2, y2=insert_y2)

        return None

@dataclass
class Table:
    bbox: BBox
    cells: List[Cell] = field(default_factory=list)
    start_page_num: Optional[int] = None
    @property
    def average_blob_height(self) -> float:
        """
        Рассчитывает среднюю высоту всех blobs во всех ячейках таблицы.
        Возвращает 0.0, если blobs отсутствуют в таблице.
        """
        all_blobs_heights = []
        for cell in self.cells:
            for blob in cell.blobs:
                all_blobs_heights.append(blob.height)
        
        if not all_blobs_heights:
            return 0.0
        return sum(all_blobs_heights) / len(all_blobs_heights)

class ParagraphType(enum.Enum):
    HEADER = 0
    FOOTER = 1
    NONE = 2

@dataclass
class Paragraph:
    bbox: BBox
    type: ParagraphType = ParagraphType.NONE
    text: str = None
    blobs: List[BBox] = field(default_factory=list)
    

@dataclass
class Page:
    image: Image.Image = None
    tables: List[Table] = field(default_factory=list)
    paragraphs: List[Paragraph] = field(default_factory=list)
    num_page: int = 0

@dataclass
class Document:
    pdf_bytes: bytes = None
    pages: List[Page] = field(default_factory=list)
    page_count: int = 0

    def get_all_text_paragraphs(self) -> str:
        '''Получем текс параграфов со всех страниц документа и представим его в виде строки'''
        full_text = []
        for page in self.pages:
            for para in page.paragraphs:
                if not para.text:
                    continue
                full_text.append(para.text)
        
        return "\n".join(full_text)

    def _get_table_column_count(self, table_obj: Table) -> int:
        """
        Calculates the number of columns in a table.
        Returns 0 if the table has no cells.
        """
        if not table_obj.cells:
            return 0
        max_col_idx = 0  # 0-indexed
        for cell in table_obj.cells:
            max_col_idx = max(max_col_idx, cell.col + cell.colspan - 1)
        return max_col_idx + 1  # Return 1-indexed count

    def get_tables(self) -> List[Table]:
        """
        Возвращает список логических таблиц из документа.
        Таблицы объединяются, если между ними (даже на разных страницах) нет ЗНАЧИМЫХ параграфов
        (колонтитулы игнорируются).
        Для объединения таблиц на разных страницах, количество их столбцов должно совпадать.
        Таблицы на одной странице всегда считаются разными логическими таблицами, если они являются
        отдельными объектами Table.
        Каждый элемент в списке - это объект Table, представляющий одну логическую таблицу.
        """
        logical_tables: List[Table] = []
        
        all_elements = []
        for page_data in self.pages:
            for p_obj in page_data.paragraphs:
                all_elements.append({
                    'type': 'paragraph', 
                    'obj': p_obj,
                    'page_num': page_data.num_page, 
                    'y1': p_obj.bbox.y1
                })
            for t_obj in page_data.tables:
                if t_obj.cells: 
                    all_elements.append({
                        'type': 'table', 
                        'obj': t_obj,
                        'page_num': page_data.num_page, 
                        'y1': t_obj.bbox.y1
                    })
        
        all_elements.sort(key=lambda x: (x['page_num'], x['y1']))

        current_accumulated_cells: List[Cell] = []
        current_row_offset: int = 0
        first_bbox_of_current_logical_table: Optional[BBox] = None
        # Page number of the first physical table fragment of the current logical table
        current_logical_table_start_page_num: Optional[int] = None 
        last_fragment_page_num: int = -1 
        current_logical_table_column_count: int = -1

        for element_data in all_elements:
            el_type = element_data['type']
            el_obj = element_data['obj']
            el_page_num = element_data['page_num']

            if el_type == 'paragraph':
                para: Paragraph = el_obj
                if para.type != ParagraphType.HEADER and para.type != ParagraphType.FOOTER:
                    if current_accumulated_cells:
                        if first_bbox_of_current_logical_table is not None:
                            logical_table = Table(
                                bbox=first_bbox_of_current_logical_table,
                                cells=list(current_accumulated_cells),
                                start_page_num=current_logical_table_start_page_num
                            )
                            logical_tables.append(logical_table)
                        
                        current_accumulated_cells.clear()
                        current_row_offset = 0
                        first_bbox_of_current_logical_table = None
                        current_logical_table_start_page_num = None
                        last_fragment_page_num = -1
                        current_logical_table_column_count = -1
            
            elif el_type == 'table':
                table_fragment: Table = el_obj
                if not table_fragment.cells:
                    continue

                fragment_column_count = self._get_table_column_count(table_fragment)

                starts_new_logical_table = False
                if not current_accumulated_cells:
                    starts_new_logical_table = True
                else:
                    if el_page_num == last_fragment_page_num:
                        starts_new_logical_table = True
                    elif el_page_num > last_fragment_page_num:
                        if (current_logical_table_column_count > 0 and
                                fragment_column_count > 0 and
                                current_logical_table_column_count != fragment_column_count):
                            starts_new_logical_table = True
                
                if starts_new_logical_table:
                    if current_accumulated_cells and first_bbox_of_current_logical_table is not None:
                        logical_table = Table(
                            bbox=first_bbox_of_current_logical_table,
                            cells=list(current_accumulated_cells),
                            start_page_num=current_logical_table_start_page_num
                        )
                        logical_tables.append(logical_table)

                    current_accumulated_cells.clear()
                    current_row_offset = 0 
                    first_bbox_of_current_logical_table = table_fragment.bbox
                    current_logical_table_start_page_num = el_page_num 
                    current_logical_table_column_count = fragment_column_count
                    
                    max_rows_in_this_fragment = 0
                    for cell in table_fragment.cells:
                        adjusted_cell = Cell(
                            bbox=cell.bbox,
                            row=cell.row + current_row_offset, 
                            col=cell.col,
                            colspan=cell.colspan,
                            rowspan=cell.rowspan,
                            text=cell.text,
                            blobs=list(cell.blobs),
                            original_page_num=el_page_num # Сохраняем исходную страницу ячейки
                        )
                        current_accumulated_cells.append(adjusted_cell)
                        max_rows_in_this_fragment = max(max_rows_in_this_fragment, cell.row + cell.rowspan)
                    
                    current_row_offset = max_rows_in_this_fragment
                    last_fragment_page_num = el_page_num

                else: 
                    max_rows_in_this_fragment = 0
                    for cell in table_fragment.cells:
                        adjusted_cell = Cell(
                            bbox=cell.bbox,
                            row=cell.row + current_row_offset, 
                            col=cell.col,
                            colspan=cell.colspan,
                            rowspan=cell.rowspan,
                            text=cell.text,
                            blobs=list(cell.blobs),
                            original_page_num=el_page_num # Сохраняем исходную страницу ячейки
                        )
                        current_accumulated_cells.append(adjusted_cell)
                        max_rows_in_this_fragment = max(max_rows_in_this_fragment, cell.row + cell.rowspan)
                    
                    current_row_offset += max_rows_in_this_fragment
                    last_fragment_page_num = el_page_num

        if current_accumulated_cells and first_bbox_of_current_logical_table is not None:
            logical_table = Table(
                bbox=first_bbox_of_current_logical_table,
                cells=list(current_accumulated_cells),
                start_page_num=current_logical_table_start_page_num
            )
            logical_tables.append(logical_table)
            
        return logical_tables

    def to_excel(self, file_path: str):
        """
        Сохраняет все логические таблицы из документа в файл Excel.
        Таблицы, разделенные параграфами, считаются новыми.
        Части таблиц без параграфов между ними (даже через страницы) объединяются.
        Каждая логическая таблица сохраняется на отдельный лист.
        Ячейки будут иметь рамки.

        Args:
            file_path (str): Путь для сохранения файла Excel.
        """
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            default_sheet = wb["Sheet"]
            wb.remove(default_sheet)

        all_elements = []
        for page_data in self.pages:
            for p_obj in page_data.paragraphs:
                all_elements.append({'type': 'paragraph', 'obj': p_obj, 
                                     'page_num': page_data.num_page, 'y1': p_obj.bbox.y1})
            for t_obj in page_data.tables:
                if t_obj.cells: 
                    all_elements.append({'type': 'table', 'obj': t_obj, 
                                         'page_num': page_data.num_page, 'y1': t_obj.bbox.y1})
        
        all_elements.sort(key=lambda x: (x['page_num'], x['y1']))

        logical_tables_cell_lists = []
        current_accumulated_cells = []
        current_row_offset = 0

        for element_data in all_elements:
            el_type = element_data['type']
            el_obj = element_data['obj']

            if el_type == 'table':
                table_fragment: Table = el_obj
                
                max_rows_in_this_fragment = 0
                if table_fragment.cells:
                    for cell in table_fragment.cells:
                        adjusted_cell = Cell(
                            bbox=cell.bbox,
                            row=cell.row + current_row_offset,
                            col=cell.col,
                            colspan=cell.colspan,
                            rowspan=cell.rowspan,
                            text=cell.text,
                            blobs=list(cell.blobs)
                        )
                        current_accumulated_cells.append(adjusted_cell)
                        max_rows_in_this_fragment = max(max_rows_in_this_fragment, cell.row + cell.rowspan)
                
                current_row_offset += max_rows_in_this_fragment
            
            elif el_type == 'paragraph':
                if current_accumulated_cells:
                    logical_tables_cell_lists.append(list(current_accumulated_cells))
                    current_accumulated_cells.clear()
                    current_row_offset = 0

        if current_accumulated_cells:
            logical_tables_cell_lists.append(list(current_accumulated_cells))

        if not logical_tables_cell_lists:
            if not wb.sheetnames:
                wb.create_sheet(title="NoTablesFound")
        else:
            # Определяем стиль границы
            thin_border_side = Side(border_style="thin", color="000000")
            cell_border = Border(left=thin_border_side, 
                                 right=thin_border_side, 
                                 top=thin_border_side, 
                                 bottom=thin_border_side)

            for i, final_table_cells in enumerate(logical_tables_cell_lists):
                sheet_name = f"Table_{i + 1}"
                if len(sheet_name) > 31: 
                    sheet_name = f"Tb{i+1}"[:31] 
                
                ws = wb.create_sheet(title=sheet_name)

                if not final_table_cells:
                    continue

                for cell_data in final_table_cells:
                    start_row_excel = cell_data.row + 1
                    start_col_excel = cell_data.col + 1
                    
                    excel_cell_obj = ws.cell(row=start_row_excel, column=start_col_excel, value=cell_data.text)
                    
                    # Применяем рамку ко всем ячейкам, которые будут частью объединенной или одиночной ячейки
                    # Для объединенных ячеек стиль применяется к верхней левой ячейке диапазона
                    excel_cell_obj.border = cell_border
                    excel_cell_obj.alignment = Alignment(wrap_text=True, vertical='top') # Выравнивание по умолчанию

                    if cell_data.rowspan > 1 or cell_data.colspan > 1:
                        end_row_excel = start_row_excel + cell_data.rowspan - 1
                        end_col_excel = start_col_excel + cell_data.colspan - 1
                        try:
                            ws.merge_cells(start_row=start_row_excel, 
                                           start_column=start_col_excel, 
                                           end_row=end_row_excel, 
                                           end_column=end_col_excel)
                            # Для объединенных ячеек, стиль рамки и выравнивание уже применены к excel_cell_obj
                            # Можно добавить специфичное выравнивание для объединенных ячеек, если нужно
                            excel_cell_obj.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        except Exception:
                            pass 
                    else: # Для одиночных ячеек рамка и выравнивание уже применены
                        excel_cell_obj.alignment = Alignment(wrap_text=True, vertical='top') # Уже установлено выше

                for col_idx_ws in range(1, ws.max_column + 1):
                    column_letter = get_column_letter(col_idx_ws)
                    ws.column_dimensions[column_letter].autosize = True
        try:
            wb.save(file_path)
        except Exception as e:
            raise

class BaseExtractor(ABC):
    def __init__(self):
        self.logger = logging.getLogger('app.' + __class__.__name__)

    def extract(self, pdf_bytes: bytes) -> Document:
        self.logger.info("Начало процесса извлечения данных из PDF.")
        if not pdf_bytes:
            self.logger.error("Получены пустые байты PDF. Прерывание операции.")
            raise ValueError("pdf_bytes не могут быть пустыми.")

        try:
            doc = pymupdf.open(stream=pdf_bytes, filetype="pdf")
        except Exception as e:
            self.logger.error(f"Ошибка при открытии PDF документа: {e}", exc_info=True)
            raise 

        page_count = len(doc)
        self.logger.info(f"Документ успешно открыт. Количество страниц: {page_count}.")
        
        pages_data: List[Page] = []
        for i in range(page_count):
            self.logger.info(f"Обработка страницы {i + 1}/{page_count}.")
            page_content = doc[i]
            try:
                paragraphs, tables = self._process(page_content)
                self.logger.debug(f"Страница {i + 1}: найдено {len(paragraphs)} параграфов и {len(tables)} таблиц.")
                
                pages_data.append(
                    Page(
                        tables=tables,
                        paragraphs=paragraphs,
                        num_page=i
                    )
                )
            except Exception as e:
                self.logger.error(f"Ошибка при обработке страницы {i + 1}: {e}", exc_info=True)

                pages_data.append(
                    Page(num_page=i) 
                )
                continue 

        self.logger.info("Все страницы обработаны. Формирование итогового документа.")
        final_document = Document(
                            pdf_bytes=pdf_bytes,
                            pages=pages_data,
                            page_count=len(pages_data)
                        )
        self.logger.info("Процесс извлечения данных из PDF завершен.")
        return final_document
    
    def _process(self, page)-> Tuple[List[Paragraph], List[Table]]:
        ...

